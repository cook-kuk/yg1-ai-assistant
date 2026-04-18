#!/usr/bin/env python3
"""Testset v5.3 — inherit v5.2 and strip the 8 ADD aggregation-sentinel rows.

Input  : testset/YG1_ARIA_Testset_v5_Shared_v5.2.xlsx
Output : testset/YG1_ARIA_Testset_v5_Shared_v5.3.xlsx

Edits:
  ADD-163/164/167/168  — clear exp_diameterMm,  drop `diameterMm=<sentinel>` token
  ADD-173/174/177/178  — clear exp_fluteCount,  drop `fluteCount=<sentinel>`  token

These rows encode sort/aggregate intent ("직경 최대값만") using filter-style
`diameterMm=최댓값만` / `fluteCount=최솟값만` sentinel values which no JSON-
schema parser can satisfy. Removing them lets the LLM score correctly on the
actual filter fields (message-level cues), and the sort intent can be
reintroduced via a future sortBy field if needed.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SRC = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared_v5.2.xlsx"
DST = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared_v5.3.xlsx"

FLUTE_IDS: set[str] = {"ADD-173", "ADD-174", "ADD-177", "ADD-178"}
DIA_IDS: set[str] = {"ADD-163", "ADD-164", "ADD-167", "ADD-168"}

STRUCTURED_COL = {
    "fluteCount": "exp_fluteCount",
    "diameterMm": "exp_diameterMm",
}
TEXT_COLS = ["필수 필터/의도", "정답지"]


def remove_token(text: str | None, field: str) -> str | None:
    """Delete a single `field=value` (or `field!=value`) token from a
    comma-separated intent string, preserving any suffix after '|'."""
    if not text:
        return text
    head, sep, tail = text.partition("|")
    parts = [p.strip() for p in head.split(",")]
    kept: list[str] = []
    changed = False
    for p in parts:
        if not p:
            changed = True
            continue
        if "=" in p:
            lhs, _, _ = p.partition("=")
            key_base = lhs.rstrip("!").rsplit(":", 1)[-1].strip()
            if key_base == field:
                changed = True
                continue
        kept.append(p)
    if not changed:
        return text
    new_head = ", ".join(kept)
    if sep:
        return f"{new_head} {sep}{tail.lstrip()}" if new_head else f"{sep}{tail.lstrip()}"
    return new_head


def main() -> int:
    if not SRC.exists():
        print(f"[fixup-v5.3] source not found: {SRC}", file=sys.stderr)
        return 1
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["🧪 테스트 시나리오"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_by_name = {name: i + 1 for i, name in enumerate(header_row) if name}
    id_col = col_by_name["질문 index"]
    flute_col = col_by_name[STRUCTURED_COL["fluteCount"]]
    dia_col = col_by_name[STRUCTURED_COL["diameterMm"]]
    text_col_idx = [col_by_name[n] for n in TEXT_COLS]

    touched: list[str] = []
    for row in ws.iter_rows(min_row=2):
        row_id = row[id_col - 1].value
        if row_id not in FLUTE_IDS and row_id not in DIA_IDS:
            continue

        if row_id in FLUTE_IDS:
            row[flute_col - 1].value = None
            for ci in text_col_idx:
                new_val = remove_token(row[ci - 1].value, "fluteCount")
                if new_val != row[ci - 1].value:
                    row[ci - 1].value = new_val
        if row_id in DIA_IDS:
            row[dia_col - 1].value = None
            for ci in text_col_idx:
                new_val = remove_token(row[ci - 1].value, "diameterMm")
                if new_val != row[ci - 1].value:
                    row[ci - 1].value = new_val
        touched.append(row_id)

    wb.save(DST)
    wb.close()

    expected = FLUTE_IDS | DIA_IDS
    missing = expected - set(touched)
    print(f"[fixup-v5.3] edited {len(touched)}/{len(expected)} rows → {DST.name}")
    if missing:
        print(f"[fixup-v5.3] WARN: not found → {sorted(missing)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
