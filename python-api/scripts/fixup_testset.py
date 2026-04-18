#!/usr/bin/env python3
"""Apply reviewed testset fixups: remove over-specified fluteCount / diameterMm
tokens from 25 ROB rows.

Input  : testset/YG1_ARIA_Testset_v5_Shared.xlsx       (unchanged)
Output : testset/YG1_ARIA_Testset_v5_Shared_v5.1.xlsx  (new copy)

Edits per row:
  - Clear the structured column (exp_fluteCount / exp_diameterMm).
  - Drop the matching `field=value` token from `필수 필터/의도` and `정답지`,
    tidy leftover dangling commas.

Safe to re-run: always overwrites the v5.1 copy from the pristine source.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SRC = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared.xlsx"
DST = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared_v5.1.xlsx"

FLUTE_IDS: set[str] = {
    "ROB-003", "ROB-005", "ROB-007", "ROB-009", "ROB-013", "ROB-015",
    "ROB-017", "ROB-019", "ROB-023", "ROB-025", "ROB-027", "ROB-029",
    "ROB-033", "ROB-035", "ROB-037", "ROB-039", "ROB-043", "ROB-045",
    "ROB-047", "ROB-049",
}
DIA_IDS: set[str] = {
    "ROB-004", "ROB-014", "ROB-024", "ROB-034", "ROB-044",
}

# Canonical field key → testset column header (exact string). Header lookup
# avoids the 0-based / 1-based indexing ambiguity between spec and python.
STRUCTURED_COL = {
    "fluteCount": "exp_fluteCount",
    "diameterMm": "exp_diameterMm",
}
TEXT_COLS = ["필수 필터/의도", "정답지"]


def remove_token(text: str | None, field: str) -> str | None:
    """Delete the single `field=value` (or `field!=value`) token from a
    comma-separated intent string. Preserves any text after '|' unchanged.
    Returns the cleaned string, or the original value if nothing matched."""
    if not text:
        return text
    head, sep, tail = text.partition("|")
    parts = [p.strip() for p in head.split(",")]
    kept: list[str] = []
    changed = False
    for p in parts:
        if not p:
            changed = True
            continue
        if "=" in p:
            # Support optional prefix like "필수 필터/의도: material"
            lhs, _, _ = p.partition("=")
            key_base = lhs.rstrip("!").rsplit(":", 1)[-1].strip()
            if key_base == field:
                changed = True
                continue
        kept.append(p)
    if not changed:
        return text
    new_head = ", ".join(kept)
    if sep:
        return f"{new_head} {sep}{tail.lstrip()}" if new_head else f"{sep}{tail.lstrip()}"
    return new_head


def main() -> int:
    if not SRC.exists():
        print(f"[fixup] source not found: {SRC}", file=sys.stderr)
        return 1

    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["🧪 테스트 시나리오"]

    # Build header → column-index map (1-based for openpyxl cell access).
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_by_name = {name: i + 1 for i, name in enumerate(header_row) if name}

    required_cols = [STRUCTURED_COL["fluteCount"], STRUCTURED_COL["diameterMm"]] + TEXT_COLS + ["질문 index"]
    for name in required_cols:
        if name not in col_by_name:
            print(f"[fixup] missing header: {name}", file=sys.stderr)
            return 2

    id_col = col_by_name["질문 index"]
    flute_col = col_by_name[STRUCTURED_COL["fluteCount"]]
    dia_col = col_by_name[STRUCTURED_COL["diameterMm"]]
    text_col_idx = [col_by_name[n] for n in TEXT_COLS]

    edited: list[tuple[str, list[str]]] = []
    for row in ws.iter_rows(min_row=2):
        row_id = row[id_col - 1].value
        if row_id not in FLUTE_IDS and row_id not in DIA_IDS:
            continue

        changes: list[str] = []
        if row_id in FLUTE_IDS:
            if row[flute_col - 1].value is not None:
                row[flute_col - 1].value = None
                changes.append("exp_fluteCount=None")
            for ci in text_col_idx:
                new_val = remove_token(row[ci - 1].value, "fluteCount")
                if new_val != row[ci - 1].value:
                    row[ci - 1].value = new_val
                    changes.append(f"col{ci}-fluteCount")
        if row_id in DIA_IDS:
            if row[dia_col - 1].value is not None:
                row[dia_col - 1].value = None
                changes.append("exp_diameterMm=None")
            for ci in text_col_idx:
                new_val = remove_token(row[ci - 1].value, "diameterMm")
                if new_val != row[ci - 1].value:
                    row[ci - 1].value = new_val
                    changes.append(f"col{ci}-diameterMm")

        edited.append((row_id, changes))

    wb.save(DST)
    wb.close()

    expected = FLUTE_IDS | DIA_IDS
    found = {rid for rid, _ in edited}
    missing = expected - found
    print(f"[fixup] edited {len(edited)}/{len(expected)} rows → {DST.name}")
    if missing:
        print(f"[fixup] WARN: not found → {sorted(missing)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
