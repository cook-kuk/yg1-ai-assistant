#!/usr/bin/env python3
"""Testset v5.2 — inherit v5.1, then fix 4 ROB rows where expected material
disagrees with the sus304 message.

Input  : testset/YG1_ARIA_Testset_v5_Shared_v5.1.xlsx   (from v5.1 fixup)
Output : testset/YG1_ARIA_Testset_v5_Shared_v5.2.xlsx   (new copy)

Edits:
  ROB-001  material=P           → material=M                         (sus304 → M)
  ROB-011  workPieceName=Alloy Steels → workPieceName=Stainless Steels
  ROB-031  workPieceName=Aluminum     → workPieceName=Stainless Steels
  ROB-041  workPieceName=Copper       → workPieceName=Stainless Steels

For each row, also updates exp_material / exp_LV1(ISO) / exp_workPieceName
columns + the same token in `필수 필터/의도` and `정답지`.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SRC = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared_v5.1.xlsx"
DST = REPO_ROOT / "testset" / "YG1_ARIA_Testset_v5_Shared_v5.2.xlsx"

# (row_id, old_token, new_token, {col_header: new_value}) — structured columns
# carry typed metadata; text columns do a literal substring replace.
EDITS: list[tuple[str, str, str, dict[str, str]]] = [
    ("ROB-001", "material=P", "material=M", {
        "exp_material": "M",
        "exp_LV1(ISO)": "M",
    }),
    ("ROB-011", "workPieceName=Alloy Steels", "workPieceName=Stainless Steels", {
        "exp_workPieceName": "Stainless Steels",
        "exp_LV1(ISO)": "M",
    }),
    ("ROB-031", "workPieceName=Aluminum", "workPieceName=Stainless Steels", {
        "exp_workPieceName": "Stainless Steels",
        "exp_LV1(ISO)": "M",
    }),
    ("ROB-041", "workPieceName=Copper", "workPieceName=Stainless Steels", {
        "exp_workPieceName": "Stainless Steels",
        "exp_LV1(ISO)": "M",
    }),
]
TEXT_COLS = ["필수 필터/의도", "정답지"]


def main() -> int:
    if not SRC.exists():
        print(f"[fixup-v5.2] source not found: {SRC}", file=sys.stderr)
        return 1
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["🧪 테스트 시나리오"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_by_name = {name: i + 1 for i, name in enumerate(header_row) if name}
    id_col = col_by_name["질문 index"]
    text_col_idx = [col_by_name[n] for n in TEXT_COLS]

    edits_by_id = {e[0]: e for e in EDITS}
    touched: list[str] = []

    for row in ws.iter_rows(min_row=2):
        row_id = row[id_col - 1].value
        if row_id not in edits_by_id:
            continue
        _, old_tok, new_tok, struct = edits_by_id[row_id]

        # 1) structured columns
        for header, val in struct.items():
            ci = col_by_name.get(header)
            if ci is None:
                print(f"[fixup-v5.2] WARN: missing header {header!r}", file=sys.stderr)
                continue
            row[ci - 1].value = val

        # 2) text columns — literal token substitution
        for ci in text_col_idx:
            cur = row[ci - 1].value or ""
            if old_tok in cur:
                row[ci - 1].value = cur.replace(old_tok, new_tok)

        touched.append(row_id)

    wb.save(DST)
    wb.close()

    expected = {e[0] for e in EDITS}
    missing = expected - set(touched)
    print(f"[fixup-v5.2] edited {len(touched)}/{len(expected)} rows → {DST.name}")
    if missing:
        print(f"[fixup-v5.2] WARN: not found → {sorted(missing)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
